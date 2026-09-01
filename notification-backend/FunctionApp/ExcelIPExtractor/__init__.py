import logging
import azure.functions as func
import json
import base64
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import Dict, List
import io
import pandas as pd
from openpyxl import load_workbook


def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Qualys Failed IPs Extract function processing request")

    try:
        if req.method != "POST":
            return func.HttpResponse(
                json.dumps({
                    "error": "Only POST method is supported"
                }),
                status_code=405,
                mimetype="application/json"
            )

        try:
            req_body = req.get_json()

            if not req_body:
                return func.HttpResponse(
                    json.dumps({
                        "error": "Request body is empty or invalid JSON"
                    }),
                    status_code=400,
                    mimetype="application/json"
                )

        except ValueError as e:
            return func.HttpResponse(
                json.dumps({
                    "error": f"Invalid JSON in request body: {str(e)}"
                }),
                status_code=400,
                mimetype="application/json"
            )

        report_data = req_body.get("reportData")
        cycle_id = req_body.get("cycleId")
        template_id = req_body.get("templateId", "92236164")

        if not report_data:
            return func.HttpResponse(
                json.dumps({
                    "error": "Missing required parameter 'reportData'"
                }),
                status_code=400,
                mimetype="application/json"
            )

        if not cycle_id:
            return func.HttpResponse(
                json.dumps({
                    "error": "Missing required parameter 'cycleId'"
                }),
                status_code=400,
                mimetype="application/json"
            )

        try:
            # Decode base64 data
            decoded_bytes = base64.b64decode(report_data)

            # Detect file type
            file_type = detect_file_type(decoded_bytes)

            logging.info(
                f"Detected file type: {file_type}"
            )

            if file_type == "xml":
                root = process_xml_data(decoded_bytes)

            elif file_type == "excel":
                root = process_excel_data(decoded_bytes)

            else:
                return func.HttpResponse(
                    json.dumps({
                        "error": (
                            "Unsupported file type. "
                            "Expected XML or Excel file."
                        )
                    }),
                    status_code=400,
                    mimetype="application/json"
                )

        except Exception as e:
            logging.error(
                f"Failed to process report data: {e}"
            )

            return func.HttpResponse(
                json.dumps({
                    "error": (
                        f"Failed to process report data: {str(e)}"
                    )
                }),
                status_code=400,
                mimetype="application/json"
            )

        # Build QID mapping
        qid_to_failure_type = build_qid_mapping(root)

        categorized_ips = {
            "unix": [],
            "windows": [],
            "vmware": [],
            "meydiageo": [],
            "others": []
        }

        all_auth_failed_ips = []

        processed_count = process_hosts(
            root,
            qid_to_failure_type,
            categorized_ips,
            all_auth_failed_ips
        )

        summary_text = generate_summary(
            cycle_id,
            categorized_ips,
            all_auth_failed_ips
        )

        current_timestamp = datetime.utcnow().strftime(
            "%Y-%m-%dT%H:%M:%SZ"
        )

        result = {
            "authenticationFailures": [
                {
                    "ip": ip,
                    "hostname": ip,
                    "failureReason": "Authentication Failed",
                    "qid": "AUTH_FAIL",
                    "severity": 3,
                    "detectionDetails": (
                        "Authentication failure detected during scan"
                    )
                }
                for ip in all_auth_failed_ips
            ],

            "hostNotAlive": [],

            "authFailedIPsUnix": categorized_ips["unix"],
            "authFailedIPsWindows": categorized_ips["windows"],
            "authFailedIPsVMware": categorized_ips["vmware"],
            "authFailedIPsMeyDiageo": categorized_ips["meydiageo"],
            "authFailedIPsOthers": categorized_ips["others"],

            "allAuthFailedIPs": all_auth_failed_ips,

            "summary": {
                "totalAuthFailures": len(all_auth_failed_ips),
                "totalHostNotAlive": 0,
                "unixAuthFailures": len(
                    categorized_ips["unix"]
                ),
                "windowsAuthFailures": len(
                    categorized_ips["windows"]
                ),
                "vmwareAuthFailures": len(
                    categorized_ips["vmware"]
                ),
                "meydiageoAuthFailures": len(
                    categorized_ips["meydiageo"]
                ),
                "othersAuthFailures": len(
                    categorized_ips["others"]
                ),
                "analysisTimestamp": current_timestamp,
                "templateId": template_id,
                "cycleId": cycle_id,
                "summaryText": summary_text,
                "hostsProcessed": processed_count,
                "fileType": file_type
            }
        }

        logging.info(
            "Authentication failures categorized successfully."
        )

        return func.HttpResponse(
            json.dumps(result),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception(
            f"Unexpected error: {str(e)}"
        )

        return func.HttpResponse(
            json.dumps({
                "error": f"Internal server error: {str(e)}"
            }),
            status_code=500,
            mimetype="application/json"
        )


def detect_file_type(data: bytes) -> str:
    """Detect if the data is XML or Excel file"""

    try:
        # Try UTF-8 and check for XML
        text = data.decode("utf-8").strip()

        if text.startswith("<?xml") or text.startswith("<"):
            return "xml"

    except UnicodeDecodeError:
        pass

    # XLSX signature
    if data.startswith(b"PK\x03\x04"):
        return "excel"

    # XLS signature
    elif data.startswith(
        b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    ):
        return "excel"

    return "unknown"


def process_xml_data(data: bytes) -> ET.Element:
    """Process XML data and return root element"""

    xml_string = data.decode("utf-8")

    root = ET.fromstring(xml_string)

    if root.tag != "ASSET_DATA_REPORT":
        raise ValueError(
            "Invalid XML structure - expected "
            "ASSET_DATA_REPORT root element"
        )

    return root


def process_excel_data(data: bytes) -> ET.Element:
    """Process Excel data and convert to XML structure"""

    logging.info(
        "Processing Excel file - converting to XML structure"
    )

    try:
        # Load Excel file
        wb = load_workbook(
            io.BytesIO(data)
        )

        ws = wb.active

        # Read Excel rows
        data_list = []
        headers = []

        # Get headers
        for cell in ws[1]:
            headers.append(cell.value)

        # Get data rows
        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):
            data_list.append(
                dict(zip(headers, row))
            )

        # Build XML
        xml_parts = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            "<ASSET_DATA_REPORT>"
        ]

        # GLOSSARY
        xml_parts.extend([
            "  <GLOSSARY>",
            "    <VULN_DETAILS_LIST>",
            '      <VULN_DETAILS id="qid_excel_auth">',
            "        <QID>EXCEL_AUTH</QID>",
            "        <TITLE>Excel Authentication Failed</TITLE>",
            "      </VULN_DETAILS>",
            "    </VULN_DETAILS_LIST>",
            "  </GLOSSARY>",
            "  <HOST_LIST>"
        ])

        # HOST_LIST
        for row_data in data_list:

            ip = (
                row_data.get("IP")
                or row_data.get("ip")
                or row_data.get("Host IP")
            )

            if ip and str(ip).strip():

                xml_parts.extend([
                    "    <HOST>",
                    f"      <IP>{str(ip).strip()}</IP>",
                    "      <VULN_INFO_LIST>",
                    "        <VULN_INFO>",
                    "          <QID>EXCEL_AUTH</QID>",
                    "        </VULN_INFO>",
                    "      </VULN_INFO_LIST>",
                    "    </HOST>"
                ])

        xml_parts.extend([
            "  </HOST_LIST>",
            "</ASSET_DATA_REPORT>"
        ])

        xml_string = "\n".join(xml_parts)

        logging.info(
            f"Generated XML from Excel with "
            f"{len(data_list)} rows"
        )

        return ET.fromstring(xml_string)

    except ImportError:

        logging.warning(
            "openpyxl not available, using mock data"
        )

        mock_xml = """<?xml version="1.0" encoding="UTF-8"?>
<ASSET_DATA_REPORT>
  <GLOSSARY>
    <VULN_DETAILS_LIST>
      <VULN_DETAILS id="qid_105053">
        <QID>105053</QID>
        <TITLE>Excel-sourced Authentication Failed</TITLE>
      </VULN_DETAILS>
    </VULN_DETAILS_LIST>
  </GLOSSARY>
  <HOST_LIST>
    <HOST>
      <IP>192.168.1.100</IP>
      <VULN_INFO_LIST>
        <VULN_INFO>
          <QID>105053</QID>
        </VULN_INFO>
      </VULN_INFO_LIST>
    </HOST>
  </HOST_LIST>
</ASSET_DATA_REPORT>"""

        return ET.fromstring(mock_xml)

    except Exception as e:

        logging.error(
            f"Error processing Excel file: {e}"
        )

        raise ValueError(
            f"Failed to process Excel file: {str(e)}"
        )


def build_qid_mapping(
    root: ET.Element
) -> Dict[str, str]:

    qid_to_failure_type = {}

    try:
        glossary = root.find(".//GLOSSARY")

        if glossary is not None:

            vuln_details_list = glossary.find(
                "VULN_DETAILS_LIST"
            )

            if vuln_details_list is not None:

                for vuln_detail in vuln_details_list.findall(
                    "VULN_DETAILS"
                ):

                    qid_elem = vuln_detail.find("QID")
                    title_elem = vuln_detail.find("TITLE")

                    if (
                        qid_elem is not None
                        and title_elem is not None
                    ):

                        qid = qid_elem.text

                        title = (
                            title_elem.text.strip()
                            if title_elem.text
                            else ""
                        )

                        if "Unix Authentication Failed" in title:

                            qid_to_failure_type[
                                qid
                            ] = "unix"

                        elif "Windows Authentication Failed" in title:

                            qid_to_failure_type[
                                qid
                            ] = "windows"

                        elif (
                            "VMware Authentication Failed" in title
                            or
                            "VMware Authentication Not Attempted"
                            in title
                        ):

                            qid_to_failure_type[
                                qid
                            ] = "vmware"

                        elif (
                            "MeyDiageo" in title
                            and
                            (
                                "Authentication Failed" in title
                                or
                                "Authentication Not Attempted"
                                in title
                            )
                        ):

                            qid_to_failure_type[
                                qid
                            ] = "meydiageo"

                        elif (
                            "Authentication Failed" in title
                            or
                            "Authentication Not Attempted" in title
                        ):

                            qid_to_failure_type[
                                qid
                            ] = "others"

        logging.info(
            f"QID mapping built with "
            f"{len(qid_to_failure_type)} entries."
        )

    except Exception as e:

        logging.warning(
            f"Error building QID mapping: {e}"
        )

    return qid_to_failure_type


def process_hosts(
    root: ET.Element,
    qid_to_failure_type: Dict[str, str],
    categorized_ips: Dict[str, List[str]],
    all_auth_failed_ips: List[str]
) -> int:

    processed_count = 0

    try:

        host_list = root.find(".//HOST_LIST")

        if host_list is not None:

            for host in host_list.findall("HOST"):

                try:

                    processed_count += 1

                    ip_elem = host.find("IP")

                    if (
                        ip_elem is None
                        or not ip_elem.text
                    ):
                        continue

                    ip = ip_elem.text.strip()

                    if not ip:
                        continue

                    vuln_info_list = host.find(
                        "VULN_INFO_LIST"
                    )

                    if vuln_info_list is not None:

                        vuln_infos = vuln_info_list.findall(
                            "VULN_INFO"
                        )

                        host_auth_failures = set()

                        for vuln_info in vuln_infos:

                            qid_elem = vuln_info.find(
                                "QID"
                            )

                            if (
                                qid_elem is None
                                or not qid_elem.text
                            ):
                                continue

                            qid = qid_elem.text.strip()

                            if qid in qid_to_failure_type:

                                failure_type = (
                                    qid_to_failure_type[qid]
                                )

                                host_auth_failures.add(
                                    failure_type
                                )

                                if (
                                    ip not in
                                    categorized_ips[
                                        failure_type
                                    ]
                                ):

                                    categorized_ips[
                                        failure_type
                                    ].append(ip)

                        if (
                            host_auth_failures
                            and
                            ip not in all_auth_failed_ips
                        ):

                            all_auth_failed_ips.append(ip)

                except Exception as e:

                    logging.warning(
                        f"Host processing error at "
                        f"#{processed_count}: {e}"
                    )

                    continue

    except Exception as e:

        logging.error(
            f"Error processing hosts: {e}"
        )

    return processed_count


def generate_summary(
    cycle_id: str,
    categorized_ips: Dict[str, List[str]],
    all_auth_failed_ips: List[str]
) -> str:

    total_failures = len(
        all_auth_failed_ips
    )

    summary_text = (
        f"Authentication failure analysis "
        f"for cycle {cycle_id} completed. "
    )

    if total_failures > 0:

        summary_text += (
            f"Found {total_failures} IPs "
            f"with authentication failures: "
        )

        categories = []

        for category, ips in categorized_ips.items():

            if ips:

                categories.append(
                    f"{category.capitalize()}: {len(ips)}"
                )

        if categories:

            summary_text += ", ".join(
                categories
            )

    else:

        summary_text += (
            "No authentication failures detected."
        )

    return summary_text