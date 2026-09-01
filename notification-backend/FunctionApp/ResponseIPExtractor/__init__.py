import logging
import azure.functions as func
import json
import base64
import re
import io
from typing import Dict, List, Union
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info(
        "Response IP Extractor function processing request"
    )

    try:
        if req.method != "POST":
            return func.HttpResponse(
                json.dumps({
                    "error": "Only POST method is supported"
                }),
                status_code=405,
                mimetype="application/json"
            )

        try:
            req_body = req.get_json()

            if not req_body:
                return func.HttpResponse(
                    json.dumps({
                        "error": "Request body is empty or invalid JSON"
                    }),
                    status_code=400,
                    mimetype="application/json"
                )

        except ValueError as e:
            return func.HttpResponse(
                json.dumps({
                    "error": f"Invalid JSON in request body: {str(e)}"
                }),
                status_code=400,
                mimetype="application/json"
            )

        data = req_body.get("data")

        if not data:
            return func.HttpResponse(
                json.dumps({
                    "error": "Missing required parameter 'data'"
                }),
                status_code=400,
                mimetype="application/json"
            )

        # Merge MeyDiageo and Diageo results
        merged_data = merge_results(data)

        # Excel export defaults to True
        export_excel = req_body.get(
            "exportExcel",
            True
        )

        # Extract IP addresses
        extraction_result = (
            extract_ip_addresses_with_categories(
                merged_data
            )
        )

        all_ips = (
            extraction_result["failedAssetIPs"]
            + extraction_result["otherIPs"]
        )

        ip_result = ",".join(all_ips)

        result = {
            "extractedIPs": ip_result,
            "ipCount": len(all_ips),
            "individualIPs": all_ips,

            "failedAssetIPs":
                extraction_result["failedAssetIPs"],

            "failedAssetIPsCount":
                len(
                    extraction_result["failedAssetIPs"]
                ),

            "otherIPs":
                extraction_result["otherIPs"],

            "otherIPsCount":
                len(
                    extraction_result["otherIPs"]
                )
        }

        # Generate Excel files if requested
        if export_excel:
            try:
                excel_data = generate_excel_reports(
                    merged_data
                )

                result["excelFiles"] = {
                    "authFailExcel":
                        excel_data["authfail_base64"],

                    "hostNotAliveExcel":
                        excel_data["hostnotalive_base64"]
                }

                logging.info(
                    "Excel files generated successfully"
                )

            except Exception as e:
                logging.error(
                    f"Error generating Excel files: {str(e)}"
                )

                result["excelError"] = str(e)

        logging.info(
            f"Successfully extracted {len(all_ips)} IP addresses "
            f"({len(extraction_result['failedAssetIPs'])} "
            f"failed assets, "
            f"{len(extraction_result['otherIPs'])} others)"
        )

        return func.HttpResponse(
            json.dumps(result),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception(
            f"Unexpected error in ResponseIPExtractor: {str(e)}"
        )

        return func.HttpResponse(
            json.dumps({
                "error": f"Internal server error: {str(e)}"
            }),
            status_code=500,
            mimetype="application/json"
        )


def merge_results(data: dict) -> dict:
    """
    Merge meyDiageoResults and diageoResults.
    """

    if not isinstance(data, dict):
        return data

    mey_results = data.get(
        "meyDiageoResults",
        {}
    )

    diageo_results = data.get(
        "diageoResults",
        {}
    )

    if not mey_results and not diageo_results:
        return data

    if not mey_results:
        return diageo_results

    if not diageo_results:
        return mey_results

    merged = {
        "failureCount":
            mey_results.get("failureCount", 0)
            + diageo_results.get("failureCount", 0),

        "failedAssets": [],

        "failedAssetsWindows": [],
        "failedAssetsUnix": [],
        "failedAssetsMeyDiageo": [],
        "failedAssetsOthers": [],

        "notAliveCount":
            mey_results.get("notAliveCount", 0)
            + diageo_results.get("notAliveCount", 0),

        "notAliveAssetsWindows": [],
        "notAliveAssetsUnix": [],
        "notAliveAssetsMeyDiageo": [],
        "notAliveAssetsUnknown": []
    }

    # Failed assets
    merged["failedAssets"].extend(
        mey_results.get("failedAssets", [])
    )

    merged["failedAssets"].extend(
        diageo_results.get("failedAssets", [])
    )

    merged["failedAssetsWindows"].extend(
        mey_results.get("failedAssetsWindows", [])
    )

    merged["failedAssetsWindows"].extend(
        diageo_results.get("failedAssetsWindows", [])
    )

    merged["failedAssetsUnix"].extend(
        mey_results.get("failedAssetsUnix", [])
    )

    merged["failedAssetsUnix"].extend(
        diageo_results.get("failedAssetsUnix", [])
    )

    merged["failedAssetsMeyDiageo"].extend(
        mey_results.get("failedAssetsMeyDiageo", [])
    )

    merged["failedAssetsMeyDiageo"].extend(
        diageo_results.get("failedAssetsMeyDiageo", [])
    )

    merged["failedAssetsOthers"].extend(
        mey_results.get("failedAssetsOthers", [])
    )

    merged["failedAssetsOthers"].extend(
        diageo_results.get("failedAssetsOthers", [])
    )

    # Not alive assets
    merged["notAliveAssetsWindows"].extend(
        mey_results.get("notAliveAssetsWindows", [])
    )

    merged["notAliveAssetsWindows"].extend(
        diageo_results.get("notAliveAssetsWindows", [])
    )

    merged["notAliveAssetsUnix"].extend(
        mey_results.get("notAliveAssetsUnix", [])
    )

    merged["notAliveAssetsUnix"].extend(
        diageo_results.get("notAliveAssetsUnix", [])
    )

    merged["notAliveAssetsMeyDiageo"].extend(
        mey_results.get("notAliveAssetsMeyDiageo", [])
    )

    merged["notAliveAssetsMeyDiageo"].extend(
        diageo_results.get("notAliveAssetsMeyDiageo", [])
    )

    merged["notAliveAssetsUnknown"].extend(
        mey_results.get("notAliveAssetsUnknown", [])
    )

    merged["notAliveAssetsUnknown"].extend(
        diageo_results.get("notAliveAssetsUnknown", [])
    )

    # Summaries
    mey_summary = mey_results.get(
        "summary",
        ""
    )

    diageo_summary = diageo_results.get(
        "summary",
        ""
    )

    if mey_summary and diageo_summary:
        merged["summary"] = (
            f"MeyDiageo: {mey_summary} | "
            f"Diageo: {diageo_summary}"
        )

    elif mey_summary:
        merged["summary"] = mey_summary

    elif diageo_summary:
        merged["summary"] = diageo_summary

    # Scan IDs
    if (
        "scanId" in mey_results
        or "scanId" in diageo_results
    ):
        merged["scanIds"] = {
            "meyDiageo":
                mey_results.get("scanId", ""),

            "diageo":
                diageo_results.get("scanId", "")
        }

    # Cycle IDs
    if (
        "cycleId" in mey_results
        or "cycleId" in diageo_results
    ):
        merged["cycleIds"] = {
            "meyDiageo":
                mey_results.get("cycleId", ""),

            "diageo":
                diageo_results.get("cycleId", "")
        }

    # Asset group names
    if (
        "assetGroupName" in mey_results
        or "assetGroupName" in diageo_results
    ):
        merged["assetGroupNames"] = {
            "meyDiageo":
                mey_results.get("assetGroupName", ""),

            "diageo":
                diageo_results.get("assetGroupName", "")
        }

    logging.info(
        f"Merged meyDiageoResults and diageoResults: "
        f"{merged['failureCount']} total failures, "
        f"{merged['notAliveCount']} total not alive"
    )

    return merged


def generate_excel_reports(
    data: Union[str, dict]
) -> Dict[str, str]:

    if isinstance(data, str):
        try:
            data = json.loads(data)

        except json.JSONDecodeError:
            data = {}

    # AuthFail Excel
    authfail_wb = create_authfail_excel(data)

    authfail_buffer = io.BytesIO()

    authfail_wb.save(authfail_buffer)

    authfail_buffer.seek(0)

    authfail_bytes = (
        authfail_buffer.getvalue()
    )

    logging.info(
        f"AuthFail Excel size: "
        f"{len(authfail_bytes)} bytes"
    )

    authfail_base64 = (
        base64.b64encode(
            authfail_bytes
        ).decode("utf-8")
    )

    # HostNotAlive Excel
    hostnotalive_wb = (
        create_hostnotalive_excel(data)
    )

    hostnotalive_buffer = io.BytesIO()

    hostnotalive_wb.save(
        hostnotalive_buffer
    )

    hostnotalive_buffer.seek(0)

    hostnotalive_bytes = (
        hostnotalive_buffer.getvalue()
    )

    logging.info(
        f"HostNotAlive Excel size: "
        f"{len(hostnotalive_bytes)} bytes"
    )

    hostnotalive_base64 = (
        base64.b64encode(
            hostnotalive_bytes
        ).decode("utf-8")
    )

    return {
        "authfail_base64":
            authfail_base64,

        "hostnotalive_base64":
            hostnotalive_base64
    }


def create_authfail_excel(
    data: dict
) -> Workbook:

    wb = Workbook()

    wb.remove(wb.active)

    sheets = {
        "Windows": [],
        "Unix-Linux": [],
        "MeyDiageo": [],
        "Others": []
    }

    failed_asset_keys = [
        "failedAssets",
        "failedAssetsWindows",
        "failedAssetsUnix",
        "failedAssetsMeyDiageo",
        "failedAssetsOthers"
    ]

    for key in failed_asset_keys:

        if (
            key in data
            and isinstance(data[key], list)
        ):

            for asset in data[key]:

                if isinstance(asset, dict):

                    ip = asset.get("ip", "")
                    hostname = asset.get(
                        "hostname",
                        ""
                    )

                    os_type = asset.get(
                        "osType",
                        ""
                    )

                    failure_reason = asset.get(
                        "failureReason",
                        ""
                    )

                    tech = asset.get(
                        "tech",
                        ""
                    )

                    category = categorize_os(
                        os_type
                    )

                    sheets[category].append({
                        "IP": ip,
                        "Hostname": hostname,
                        "OS": os_type,
                        "Failure Reason":
                            failure_reason,
                        "Tech": tech
                    })

    for sheet_name, assets in sheets.items():

        ws = wb.create_sheet(
            title=sheet_name
        )

        create_authfail_sheet(
            ws,
            assets
        )

    return wb


def create_authfail_sheet(
    ws,
    assets: List[dict]
):

    headers = [
        "IP",
        "Hostname",
        "OS",
        "Failure Reason",
        "Tech"
    ]

    header_fill = PatternFill(
        start_color="366092",
        end_color="366092",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for col_idx, header in enumerate(
        headers,
        1
    ):

        cell = ws.cell(
            row=1,
            column=col_idx,
            value=header
        )

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    unique_assets = []
    seen = set()

    for asset in assets:

        asset_key = (
            asset.get("IP", ""),
            asset.get("Hostname", ""),
            asset.get("OS", ""),
            asset.get("Failure Reason", ""),
            asset.get("Tech", "")
        )

        if asset_key not in seen:

            seen.add(asset_key)
            unique_assets.append(asset)

    for row_idx, asset in enumerate(
        unique_assets,
        2
    ):

        ws.cell(
            row=row_idx,
            column=1,
            value=asset.get("IP", "")
        )

        ws.cell(
            row=row_idx,
            column=2,
            value=asset.get("Hostname", "")
        )

        ws.cell(
            row=row_idx,
            column=3,
            value=asset.get("OS", "")
        )

        ws.cell(
            row=row_idx,
            column=4,
            value=asset.get(
                "Failure Reason",
                ""
            )
        )

        ws.cell(
            row=row_idx,
            column=5,
            value=asset.get("Tech", "")
        )

    for col_idx in range(1, 6):
        ws.column_dimensions[
            chr(64 + col_idx)
        ].width = 25


def create_hostnotalive_excel(
    data: dict
) -> Workbook:

    wb = Workbook()

    wb.remove(wb.active)

    sheet_mapping = {
        "Windows":
            "notAliveAssetsWindows",

        "Unix-Linux":
            "notAliveAssetsUnix",

        "MeyDiageo":
            "notAliveAssetsMeyDiageo",

        "Unknown":
            "notAliveAssetsUnknown"
    }

    for sheet_name, data_key in (
        sheet_mapping.items()
    ):

        ws = wb.create_sheet(
            title=sheet_name
        )

        ips = data.get(
            data_key,
            []
        )

        create_hostnotalive_sheet(
            ws,
            ips
        )

    return wb


def create_hostnotalive_sheet(
    ws,
    ips: List[str]
):

    header_fill = PatternFill(
        start_color="366092",
        end_color="366092",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    cell = ws.cell(
        row=1,
        column=1,
        value="IP"
    )

    cell.fill = header_fill
    cell.font = header_font

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    unique_ips = []
    seen = set()

    for ip in ips:

        if ip and ip not in seen:

            seen.add(ip)
            unique_ips.append(ip)

    for row_idx, ip in enumerate(
        unique_ips,
        2
    ):

        ws.cell(
            row=row_idx,
            column=1,
            value=ip
        )

    ws.column_dimensions["A"].width = 20


def categorize_os(
    os_type: str
) -> str:

    if not os_type:
        return "Others"

    os_lower = os_type.lower()

    if "windows" in os_lower:
        return "Windows"

    elif (
        "linux" in os_lower
        or "unix" in os_lower
    ):
        return "Unix-Linux"

    elif (
        "meydiageo" in os_lower
        or "mey diageo" in os_lower
    ):
        return "MeyDiageo"

    else:
        return "Others"


def extract_ip_addresses_with_categories(
    data: Union[str, dict]
) -> dict:

    failed_asset_ips = []
    other_ips = []

    try:

        if isinstance(data, str):

            try:
                data = json.loads(data)

            except json.JSONDecodeError:

                other_ips = extract_ips_from_text(
                    data
                )

                return {
                    "failedAssetIPs":
                        [],
                    "otherIPs":
                        remove_duplicates(
                            other_ips
                        )
                }

        if isinstance(data, dict):

            (
                failed_asset_ips,
                other_ips
            ) = extract_ips_from_json_categorized(
                data
            )

        else:

            other_ips = extract_ips_from_text(
                str(data)
            )

        unique_failed_ips = (
            remove_duplicates(
                failed_asset_ips
            )
        )

        unique_other_ips = (
            remove_duplicates(
                other_ips
            )
        )

        unique_other_ips = [
            ip
            for ip in unique_other_ips
            if ip not in unique_failed_ips
        ]

        return {
            "failedAssetIPs":
                unique_failed_ips,

            "otherIPs":
                unique_other_ips
        }

    except Exception as e:

        logging.error(
            f"Error extracting IP addresses: {e}"
        )

        return {
            "failedAssetIPs": [],
            "otherIPs": []
        }


def extract_ips_from_json_categorized(
    data: dict
) -> tuple:

    failed_asset_ips = []
    other_ips = []

    failed_asset_keys = [
        "failedAssets",
        "failedAssetsWindows",
        "failedAssetsUnix",
        "failedAssetsMeyDiageo",
        "failedAssetsOthers"
    ]

    def extract_from_failed_assets(value):

        if isinstance(value, list):

            for item in value:

                if isinstance(item, dict):

                    if (
                        "ip" in item
                        and isinstance(
                            item["ip"],
                            str
                        )
                    ):

                        if is_valid_ip(
                            item["ip"]
                        ):

                            failed_asset_ips.append(
                                item["ip"]
                            )

                    for field_value in item.values():

                        if (
                            isinstance(
                                field_value,
                                str
                            )
                            and
                            field_value
                            != item.get(
                                "ip",
                                ""
                            )
                        ):

                            ip_pattern = (
                                r"\b(?:[0-9]{1,3}\.){3}"
                                r"[0-9]{1,3}\b"
                            )

                            matches = re.findall(
                                ip_pattern,
                                field_value
                            )

                            for match in matches:

                                if is_valid_ip(
                                    match
                                ):

                                    failed_asset_ips.append(
                                        match
                                    )

    def extract_from_other_sources(
        value,
        key_name=""
    ):

        if isinstance(value, str):

            if is_valid_ip(value):

                other_ips.append(value)

            else:

                ip_pattern = (
                    r"\b(?:[0-9]{1,3}\.){3}"
                    r"[0-9]{1,3}\b"
                )

                matches = re.findall(
                    ip_pattern,
                    value
                )

                for match in matches:

                    if is_valid_ip(match):

                        other_ips.append(match)

        elif isinstance(value, dict):

            for k, v in value.items():

                extract_from_other_sources(
                    v,
                    k
                )

        elif isinstance(value, list):

            for item in value:

                if (
                    isinstance(item, str)
                    and is_valid_ip(item)
                ):

                    other_ips.append(item)

                else:

                    extract_from_other_sources(
                        item,
                        key_name
                    )

    for key, value in data.items():

        if key in failed_asset_keys:

            extract_from_failed_assets(
                value
            )

        else:

            extract_from_other_sources(
                value,
                key
            )

    return (
        failed_asset_ips,
        other_ips
    )


def extract_ips_from_text(
    data: str
) -> List[str]:

    ip_addresses = []

    try:

        lines = data.strip().split("\n")

        ip_pattern = (
            r"\b(?:[0-9]{1,3}\.){3}"
            r"[0-9]{1,3}\b"
        )

        for line in lines:

            line = line.strip()

            if line:

                matches = re.findall(
                    ip_pattern,
                    line
                )

                for match in matches:

                    if is_valid_ip(match):

                        ip_addresses.append(
                            match
                        )

        return ip_addresses

    except Exception as e:

        logging.error(
            f"Error extracting IP addresses from text: {e}"
        )

        return []


def remove_duplicates(
    ip_list: List[str]
) -> List[str]:

    unique_ips = []

    for ip in ip_list:

        if ip not in unique_ips:

            unique_ips.append(ip)

    return unique_ips


def is_valid_ip(
    ip: str
) -> bool:

    try:

        parts = ip.split(".")

        if len(parts) != 4:
            return False

        for part in parts:

            num = int(part)

            if num < 0 or num > 255:
                return False

        return True

    except (
        ValueError,
        AttributeError
    ):

        return False